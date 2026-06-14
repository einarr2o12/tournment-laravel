#!/usr/bin/env python3
# Run from project root on the HOST (needs openpyxl):  python3 reports/build_xlsx.py
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
BRAND, BRAND_DK, LIGHT, GREY = "4F46E5", "312E81", "EEF2FF", "F1F5F9"
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(size=16, bold=True, color=BRAND_DK)
sub_font = Font(size=10, color="64748B")
hdr_font = Font(size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=BRAND)
group_fill = PatternFill("solid", fgColor=LIGHT)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")


def header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, center, border


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load(name):
    return json.load(open(os.path.join(HERE, name)))


# ---- GROUPS ----
g = load("groups_data.json")
wb = Workbook(); ws = wb.active; ws.title = "Groups"
ws.merge_cells("A1:F1"); ws["A1"] = g["tournament"]; ws["A1"].font = title_font
ws.merge_cells("A2:F2"); ws["A2"] = f"Group Stage Draw  ·  {g.get('venue','')}  ·  Jun 21, 2026"; ws["A2"].font = sub_font
for c, h in enumerate(["Category", "Group", "Seed", "Team", "Player 1", "Player 2"], 1):
    ws.cell(row=4, column=c, value=h)
header(ws, 4, 6)
r, pc, pg = 5, None, None
for row in g["rows"]:
    nb = (row["category"], row["group"]) != (pc, pg)
    vals = [row["category"] if nb else "", row["group"] if nb else "", row["seed"], row["team"], row["player1"], row["player2"]]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v); cell.border = border
        cell.alignment = center if c in (2, 3) else left
        if nb and c <= 2:
            cell.fill = group_fill; cell.font = Font(bold=True, color=BRAND_DK)
    pc, pg = row["category"], row["group"]; r += 1
widths(ws, [34, 9, 7, 28, 20, 20]); ws.freeze_panes = "A5"
wb.save(os.path.join(HERE, "Groups_Report.xlsx"))

# ---- SCHEDULE ----
s = load("schedule_data.json")
wb2 = Workbook(); w1 = wb2.active; w1.title = "By Time"
w1.merge_cells("A1:H1"); w1["A1"] = s["tournament"]; w1["A1"].font = title_font
w1.merge_cells("A2:H2"); w1["A2"] = f"Group Stage Schedule  ·  {s.get('venue','')}  ·  Jun 21, 2026  ·  {len(s['rows'])} matches (each 1 game to 21)"; w1["A2"].font = sub_font
for c, h in enumerate(["Time", "Court", "Category", "Group", "Round", "Team A", "Team B", "Status"], 1):
    w1.cell(row=4, column=c, value=h)
header(w1, 4, 8)
r, pt = 5, None
for row in s["rows"]:
    nb = row["time"] != pt
    vals = [row["time"] if nb else "", row["court"], row["category"], row["group"], f"R{row['round']}", row["teamA"], row["teamB"], row["status"]]
    for c, v in enumerate(vals, 1):
        cell = w1.cell(row=r, column=c, value=v); cell.border = border
        cell.alignment = center if c in (1, 2, 4, 5, 8) else left
        if nb:
            cell.fill = group_fill if c == 1 else PatternFill("solid", fgColor=GREY)
            if c == 1: cell.font = Font(bold=True, color=BRAND_DK)
    pt = row["time"]; r += 1
widths(w1, [11, 9, 34, 9, 7, 26, 26, 12]); w1.freeze_panes = "A5"

# By Court matrix
w2 = wb2.create_sheet("By Court")
times = sorted({r["time"] for r in s["rows"]}, key=lambda t: datetime.strptime(t, "%I:%M %p"))
courts = sorted({r["court"] for r in s["rows"]}, key=lambda c: int(c.split()[-1]) if c.split()[-1].isdigit() else 99)
w2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(courts) + 1)
w2["A1"] = "Schedule by Court  ·  Jun 21, 2026"; w2["A1"].font = title_font
w2.cell(row=3, column=1, value="Time")
for i, ct in enumerate(courts, 2): w2.cell(row=3, column=i, value=ct)
header(w2, 3, len(courts) + 1)
grid = {(r["time"], r["court"]): f"{r['teamA']}  vs  {r['teamB']}\n({r['category'][:18]} {r['group'][-1]})" for r in s["rows"]}
for ri, tm in enumerate(times, 4):
    tc = w2.cell(row=ri, column=1, value=tm); tc.font = Font(bold=True, color=BRAND_DK); tc.fill = group_fill; tc.border = border; tc.alignment = center
    for ci, ct in enumerate(courts, 2):
        cell = w2.cell(row=ri, column=ci, value=grid.get((tm, ct), "")); cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    w2.row_dimensions[ri].height = 34
widths(w2, [11] + [30] * len(courts)); w2.freeze_panes = "B4"
wb2.save(os.path.join(HERE, "Schedule_Report.xlsx"))
print("built reports/Groups_Report.xlsx + reports/Schedule_Report.xlsx")
